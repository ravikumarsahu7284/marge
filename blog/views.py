from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from .models import Post, Category, Tags, Comment, User
from .forms import UserForm, LoginForm, PostForm, CommentForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
from .models import Hoteldata


#register form
def register(request):
    if request.method == "POST":
        form = UserForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False) # <---
            user.set_password(form.cleaned_data.get('password'))             
            # form.set_password   
            form.save()
            return redirect("login")
    else:
        form = UserForm()
    return render(request, "blog/register.html", {"form": form})

#login form
def login_page(request):
    form = LoginForm()
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
            )
            if user:
                login(request, user)
                return redirect("post_list")
    else:
        form = LoginForm()
    return render(request, 'blog/login.html', context={'form': form})

#user profile view
def profile(request):
        profile = request.user
        return render(request, 'blog/profile.html', {'profile': profile})

#profile edit
@login_required
def edit_profile(request):
    profile = request.user
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = UserForm(instance=profile)
    return render(request, 'blog/edit_profile.html', {'form': form})

#logout page
def logout_user(request):
    logout(request)
    return redirect('login')

# Create post list and home page post show .
def post_list(request):
    posts = Post.objects.all()
    return render(request, 'blog/post_list.html', {'posts': posts}) 

# create post category filter
def post_category(request, slug):
    categories = Category.objects.filter(slug=slug).last()
    posts = Post.objects.filter(category=categories).all()
    # context = {'category': category, 'posts': posts}
    return render(request, 'blog/post_category.html', {'posts': posts}) 

# create post tags filter
def post_tags(request, slug):
    tag = Tags.objects.filter(slug=slug).last()
    posts = Post.objects.filter(tags=tag).all()
    # print(posts, 'ssssssssssssssssss')
    # context = {'category': category, 'posts': posts}
    return render(request, 'blog/post_tags.html', {'posts': posts}) 

# create post autor filter 
def post_author(request, slug):
    authors = User.objects.filter(username=slug).last()
    # print(authors, 'ssssssssssssss')
    # author = request.post.author
    posts = Post.objects.filter(author=authors).all()
    # print(posts, 'rrrrrrrrrrrrr')
    return render(request, 'blog/authorfilter.html', {'posts': posts}) 

def post_date(request, slug):
    published = Post.objects.filter(published_date=slug).last()
    # print(authors, 'ssssssssssssss')
    # author = request.post.author
    posts = Post.objects.filter(published_date=published).all()
    # print(posts, 'rrrrrrrrrrrrr')
    return render(request, 'blog/postdatefilter.html', {'posts': posts}) 

#Post page and comment
def post_detail(request, slug):
    post = get_object_or_404(Post, slug=slug)
    comments = Comment.objects.filter(post=post).order_by('-id')
    form = CommentForm()
    new_comment = None
    reply = None
    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            try:
                reply = request.POST.get('comment_id')
                reply = Comment.objects.filter(id=reply).last()
            except:
                reply=None         
        new_comment = form.save(commit=False)
        new_comment.reply = reply
        new_comment.post = post
        new_comment.save()  
        return redirect('post_detail', slug=post.slug)
    else:
        form = CommentForm()   
        return render(request, 'blog/post_detail.html', {'post':post,'comments':comments,'form': form, 'reply':reply})

# create new post 
def post_new(request):
    form = PostForm()
    if request.method == "POST":
        form = PostForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            post.published_date = timezone.now()    
            post.save()
            return redirect('post_detail', slug=post.slug)
    else:
        form = PostForm()   
    return render(request, 'blog/post_edit.html', {'form': form})  

# create post edit 
def post_edit(request, slug):
    post = get_object_or_404(Post, slug=slug)
    if request.method == "POST":
        form = PostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            post.published_date = timezone.now()
            post.tags.set(form.cleaned_data.get('tags')) 
            post.save()
            return redirect('post_detail', slug=post.slug)
    else:
        form = PostForm(instance=post)
        
    return render(request, 'blog/post_edit.html', {'form': form})

# create a excel file upload and append the data in same page with file handling 
def upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            exceldata = load_workbook(excel_file, data_only=True)
            ws = exceldata.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                continent, country, city, hotelname, stars, date, end_date, price, discounted_price = row
                # print(continent, country, city, hotelname, stars, date, end_date, price, discounted_price, '9999999999999999999999999999')

                # Convert strings to appropriate data types
                date = datetime.strptime(date, '%d/%m/%Y')
                end_date = datetime.strptime(end_date, '%d/%m/%Y')

                # If price is already an integer, use it directly
                if isinstance(price, int):
                    price_decimal = Decimal(price)
                else:
                    price_decimal = Decimal(price.replace('.', '').replace(',', '.'))

                if isinstance(discounted_price, int):
                    discounted_price_decimal = Decimal(discounted_price)
                else:
                    discounted_price_decimal = Decimal(discounted_price.replace('.', '').replace(',', '.'))

                if Hoteldata.objects.filter(continent=continent, country=country, city=city, hotelname=hotelname, stars=stars, date=date, end_date=end_date, price=price_decimal, discounted_Price=discounted_price_decimal).exists():
                    pass
                else :
                    Hoteldata.objects.create(continent=continent, country=country,
                    city=city, hotelname=hotelname, stars=stars, date=date, end_date=end_date, price=price_decimal, discounted_Price=discounted_price_decimal)
                
    hoteldata_list = Hoteldata.objects.all().order_by('-id')
    return render(request, 'blog/task.html', {'hoteldata_list': hoteldata_list})
